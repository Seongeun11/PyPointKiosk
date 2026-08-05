import os
import json
from datetime import datetime
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ReceiptRepositoryModel:
    """영수증 JSON 저장(1~999) 및 엑셀 보고서 변환 비즈니스 Model"""
    
    def __init__(self, json_path="resources/receipts.json", products_path="resources/products.json"):
        self.base_dir = os.path.dirname(os.path.dirname(__file__))
        self.json_path = os.path.join(self.base_dir, json_path)
        self.products_path = os.path.join(self.base_dir, products_path)
        self._ensure_file_exists()

    def _ensure_file_exists(self):
        os.makedirs(os.path.dirname(self.json_path), exist_ok=True)
        if not os.path.exists(self.json_path):
            with open(self.json_path, "w", encoding="utf-8") as f:
                json.dump([], f, ensure_ascii=False, indent=2)

    def add_receipt(self, pay_type: str, cart_items: list, purchase_amount: int, 
                    discount_type: str, discount_amount: int, final_amount: int) -> dict:
        """결제 완료 시 영수증 데이터를 JSON 배열(1~999 순번)에 추가"""
        try:
            with open(self.json_path, "r", encoding="utf-8") as f:
                receipts = json.load(f)
        except Exception:
            receipts = []

        next_id = (len(receipts) % 999) + 1

        receipt_data = {
            "id": next_id,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "pay_type": pay_type,                     
            "discount_type": discount_type,           
            "purchase_amount": purchase_amount,
            "discount_amount": discount_amount,
            "final_amount": final_amount,
            "items": cart_items                       
        }

        receipts.append(receipt_data)

        with open(self.json_path, "w", encoding="utf-8") as f:
            json.dump(receipts, f, ensure_ascii=False, indent=2)

        return receipt_data

    def get_all_receipts(self) -> list:
        try:
            with open(self.json_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []

    def export_to_excel(self, export_file_path: str) -> bool:
        """JSON 영수증 집계 데이터를 일일 판매 보고서 양식 엑셀 파일로 출력"""
        receipts = self.get_all_receipts()
        
        products_data = []
        if os.path.exists(self.products_path):
            with open(self.products_path, "r", encoding="utf-8") as f:
                cat_data = json.load(f).get("categories", [])
                for cat in cat_data:
                    c_name = cat.get("name", "")
                    for p in cat.get("products", []):
                        products_data.append({
                            "category": c_name,
                            "name": p.get("name", ""),
                            "price": p.get("price", 0)
                        })

        wb = openpyxl.Workbook()
        ws = wb.active

        # Pylance 타입 에러 방지 (Type Guard)
        if ws is None or not isinstance(ws, Worksheet):
            ws = wb.create_sheet(title="일일 판매 보고서")
        else:
            ws.title = "일일 판매 보고서"

        # 스타일 정의
        font_title = Font(name="맑은 고딕", size=14, bold=True)
        font_header = Font(name="맑은 고딕", size=9, bold=True)
        
        fill_title = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_sub_title = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center')

        # 1. 헤더 생성 (1~3행)
        ws.merge_cells("A1:N2")
        ws["A1"].value = "만물복귀 카페팀 일일 판매 보고서"
        ws["A1"].font = font_title
        ws["A1"].fill = fill_title
        ws["A1"].alignment = align_center

        ws.merge_cells("O1:R2")
        ws["O1"].value = datetime.now().strftime("%Y. %m. %d")
        ws["O1"].font = font_title
        ws["O1"].fill = fill_title
        ws["O1"].alignment = align_center

        headers_row3 = [
            ("A3", "품목"), ("B3", "가격표"), ("F3", "할인가"), ("H3", "일반가"), 
            ("J3", "아카데미"), ("L3", "엔화"), ("M3", "엔화 합계"), 
            ("N3", "현금 합계"), ("O3", "계좌 합계"), ("P3", "총 합계"), 
            ("Q3", "아카데미 포인트"), ("R3", "총 판매량")
        ]
        
        ws.merge_cells("B3:E3")
        ws.merge_cells("F3:G3")
        ws.merge_cells("H3:I3")
        ws.merge_cells("J3:K3")

        for cell_ref, text in headers_row3:
            ws[cell_ref].value = text
            ws[cell_ref].font = font_header
            ws[cell_ref].alignment = align_center
            ws[cell_ref].fill = fill_sub_title

        # 서브 헤더 (4행)
        sub_headers = {
            "B4": "할인", "C4": "원", "D4": "아카데미", "E4": "엔",
            "F4": "현금", "G4": "계좌", "H4": "현금", "I4": "계좌",
            "J4": "현금", "K4": "계좌"
        }
        for cell_ref, text in sub_headers.items():
            ws[cell_ref].value = text
            ws[cell_ref].font = font_header
            ws[cell_ref].alignment = align_center
            ws[cell_ref].fill = fill_sub_title

        # 2. 데이터 집계 계산
        stats = {}
        for p in products_data:
            p_name = p["name"]
            stats[p_name] = {
                "category": p["category"],
                "price": p["price"],
                "disc_cash": 0, "disc_bank": 0,
                "norm_cash": 0, "norm_bank": 0,
                "acad_cash": 0, "acad_bank": 0,
                "point_qty": 0,
                "total_qty": 0
            }

        for r in receipts:
            p_type = r.get("pay_type")
            d_type = r.get("discount_type")
            items = r.get("items", [])

            for item in items:
                name = item.get("name")
                qty = item.get("quantity", 0)

                if name not in stats:
                    stats[name] = {
                        "category": "기타", "price": item.get("price", 0),
                        "disc_cash": 0, "disc_bank": 0,
                        "norm_cash": 0, "norm_bank": 0,
                        "acad_cash": 0, "acad_bank": 0,
                        "point_qty": 0, "total_qty": 0
                    }

                st = stats[name]
                st["total_qty"] += qty

                if p_type == "point":
                    st["point_qty"] += qty
                elif d_type == "student":
                    if p_type == "cash": st["disc_cash"] += qty
                    elif p_type == "bank": st["disc_bank"] += qty
                elif d_type == "academy":
                    if p_type == "cash": st["acad_cash"] += qty
                    elif p_type == "bank": st["acad_bank"] += qty
                else:
                    if p_type == "cash": st["norm_cash"] += qty
                    elif p_type == "bank": st["norm_bank"] += qty

        # 3. 데이터 엑셀 행 채우기 (5행부터)
        row_idx = 5
        for name, data in stats.items():
            price = data["price"]
            disc_price = int(price * 0.9)
            acad_price = int(price * 0.85)

            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=f"₩{disc_price:,}")
            ws.cell(row=row_idx, column=3, value=f"₩{price:,}")
            ws.cell(row=row_idx, column=4, value=f"₩{acad_price:,}")
            ws.cell(row=row_idx, column=5, value=f"¥{int(price/10)}")

            ws.cell(row=row_idx, column=6, value=data["disc_cash"] or "")
            ws.cell(row=row_idx, column=7, value=data["disc_bank"] or "")
            ws.cell(row=row_idx, column=8, value=data["norm_cash"] or "")
            ws.cell(row=row_idx, column=9, value=data["norm_bank"] or "")
            ws.cell(row=row_idx, column=10, value=data["acad_cash"] or "")
            ws.cell(row=row_idx, column=11, value=data["acad_bank"] or "")
            ws.cell(row=row_idx, column=12, value="")

            ws.cell(row=row_idx, column=13, value=f"=L{row_idx}*E{row_idx}")
            ws.cell(row=row_idx, column=14, value=f"=(F{row_idx}*{disc_price})+(H{row_idx}*{price})+(J{row_idx}*{acad_price})")
            ws.cell(row=row_idx, column=15, value=f"=(G{row_idx}*{disc_price})+(I{row_idx}*{price})+(K{row_idx}*{acad_price})")
            ws.cell(row=row_idx, column=16, value=f"=N{row_idx}+O{row_idx}")

            ws.cell(row=row_idx, column=17, value=data["point_qty"] or "")
            ws.cell(row=row_idx, column=18, value=f"=SUM(F{row_idx}:L{row_idx})+Q{row_idx}")

            row_idx += 1

        # 전체 테두리 적용 및 정렬
        for r in range(1, row_idx):
            for c in range(1, 19):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                if r >= 5 and c not in [1, 2, 3, 4, 5]:
                    cell.alignment = align_center if c <= 12 or c in [17, 18] else align_right

        wb.save(export_file_path)
        return True