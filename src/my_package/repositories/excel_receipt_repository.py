import os
import json
from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ReceiptRepositoryModel:
    """
    영수증 일자별 JSON 저장 및 엑셀 보고서 변환 비즈니스 Model
    수련생/아카데미 고정 할인액 정산 및 엑셀 연동 기능 지원
    """
    CLOSING_OFFSET_HOURS = 2

    def __init__(self, base_receipts_dir: str = "resources/receipts", products_path: str = "resources/products.json"):
        self.base_dir = os.path.dirname(os.path.dirname(__file__))
        self.receipts_dir = os.path.join(self.base_dir, base_receipts_dir)
        self.products_path = os.path.join(self.base_dir, products_path)
        os.makedirs(self.receipts_dir, exist_ok=True)

    def _get_business_date_str(self, dt: Optional[datetime] = None) -> str:
        if dt is None:
            dt = datetime.now()
        business_dt = dt - timedelta(hours=self.CLOSING_OFFSET_HOURS)
        return business_dt.strftime("%Y-%m-%d")

    def _get_daily_file_path(self, date_str: Optional[str] = None) -> str:
        if not date_str:
            date_str = self._get_business_date_str()
        filename = f"receipts_{date_str}.json"
        return os.path.join(self.receipts_dir, filename)

    def _load_receipts_by_path(self, file_path: str) -> List[Dict[str, Any]]:
        if not os.path.exists(file_path):
            return []
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"[Model Error] 영수증 파일 읽기 실패 ({file_path}): {e}")
            return []

    def add_receipt(self, pay_type: str, cart_items: list, purchase_amount: int, 
                    discount_type: str, discount_amount: int, final_amount: int,
                    currency: str = "KRW") -> dict:
        now = datetime.now()
        business_date_str = self._get_business_date_str(now)
        daily_file_path = self._get_daily_file_path(business_date_str)

        receipts = self._load_receipts_by_path(daily_file_path)
        next_id = (len(receipts) % 999) + 1

        receipt_data = {
            "id": next_id,
            "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
            "pay_type": pay_type,                     
            "discount_type": discount_type,           
            "currency": currency,
            "purchase_amount": purchase_amount,
            "discount_amount": discount_amount,       
            "final_amount": final_amount,             
            "items": cart_items                       
        }

        receipts.append(receipt_data)

        try:
            with open(daily_file_path, "w", encoding="utf-8") as f:
                json.dump(receipts, f, ensure_ascii=False, indent=2)
                print(f"[Model] 영수증 저장 완료 (통화: {currency}, 영업일: {business_date_str}): {daily_file_path}")
        except Exception as e:
            print(f"[Model Error] 영수증 저장 실패: {e}")

        return receipt_data

    def get_receipts_by_date(self, date_str: Optional[str] = None) -> list:
        file_path = self._get_daily_file_path(date_str)
        return self._load_receipts_by_path(file_path)

    def export_to_excel(self, export_file_path: str, target_date: Optional[str] = None) -> bool:
        """
        [수련생/아카데미/커스텀 할인 세분화 정산 엑셀 생성]
        - 가격표: 정가, 수련생 할인액, 아카데미 할인액 명시
        - 매출 수식: 가격표의 정가 기준 소계 산출 후, 하단 차감 행에서 실제 할인액 차감
        """
        if not target_date:
            target_date = self._get_business_date_str()

        default_dir = os.path.join(self.base_dir, "resources/reports")
        os.makedirs(default_dir, exist_ok=True)

        if not export_file_path:
            export_file_path = os.path.join(default_dir, f"일일_판매_보고서_{target_date}.xlsx")
        else:
            if not export_file_path.endswith(".xlsx"):
                os.makedirs(export_file_path, exist_ok=True)
                export_file_path = os.path.join(export_file_path, f"일일_판매_보고서_{target_date}.xlsx")
            else:
                base, ext = os.path.splitext(export_file_path)
                if not base.endswith(target_date):
                    export_file_path = f"{base}_{target_date}{ext}"

        receipts = self.get_receipts_by_date(target_date)
        
        # 1. Master 상품 목록 구조화 (products.json의 수련생/아카데미 할인액 읽기)
        stats: Dict[str, Dict[str, Any]] = {}
        if os.path.exists(self.products_path):
            try:
                with open(self.products_path, "r", encoding="utf-8") as f:
                    cat_data = json.load(f).get("categories", [])
                    for cat in cat_data:
                        c_name = cat.get("name", "미지정")
                        for p in cat.get("products", []):
                            prod_id = str(p.get("id"))
                            stats[prod_id] = {
                                "id": prod_id,
                                "name": p.get("name", ""),
                                "category": c_name,
                                "price": p.get("price", 0),
                                "discount_student": p.get("discount_student", 0), # 수련생 할인액
                                "discount_academy": p.get("discount_academy", 0), # 아카데미 할인액
                                "disc_cash": 0, "disc_bank": 0, "disc_jpy": 0,
                                "norm_cash": 0, "norm_bank": 0, "norm_jpy": 0,
                                "acad_cash": 0, "acad_bank": 0,
                                "point_qty": 0, "total_qty": 0
                            }
            except Exception as e:
                print(f"[Model Error] 상품 목록 로드 실패: {e}")

        # 할인 금액 종류별 집계 변수 (KRW / JPY)
        discount_totals = {
            "student_cash_krw": 0, "student_bank_krw": 0, "student_jpy": 0,
            "academy_cash_krw": 0, "academy_bank_krw": 0, "academy_jpy": 0,
            "custom_cash_krw": 0,  "custom_bank_krw": 0,  "custom_jpy": 0
        }

        # 2. 영수증 JSON 내역 상세 집계
        for r in receipts:
            p_type = r.get("pay_type")        
            d_type = r.get("discount_type")   
            currency = r.get("currency", "KRW")
            r_discount_amt = r.get("discount_amount", 0)
            items = r.get("items", [])

            # 할인 유형 및 결제 수단별 누적 차감액 집계
            if d_type == "student":
                if currency == "JPY": 
                    discount_totals["student_jpy"] += r_discount_amt
                elif p_type == "bank": 
                    discount_totals["student_bank_krw"] += r_discount_amt
                else: 
                    discount_totals["student_cash_krw"] += r_discount_amt

            elif d_type == "academy":
                if currency == "JPY": 
                    discount_totals["academy_jpy"] += r_discount_amt
                elif p_type == "bank": 
                    discount_totals["academy_bank_krw"] += r_discount_amt
                else: 
                    discount_totals["academy_cash_krw"] += r_discount_amt

            elif r_discount_amt > 0:
                if currency == "JPY": 
                    discount_totals["custom_jpy"] += r_discount_amt
                elif p_type == "bank": 
                    discount_totals["custom_bank_krw"] += r_discount_amt
                else: 
                    discount_totals["custom_cash_krw"] += r_discount_amt

            for item in items:
                prod_id = str(item.get("id", item.get("name")))
                name = item.get("name", "미등록상품")
                qty = item.get("quantity", 0)
                item_currency = item.get("currency", currency)

                if prod_id not in stats:
                    stats[prod_id] = {
                        "id": prod_id,
                        "name": name,
                        "category": "기타", 
                        "price": item.get("price", 0),
                        "discount_student": item.get("discount_student", 0),
                        "discount_academy": item.get("discount_academy", 0),
                        "disc_cash": 0, "disc_bank": 0, "disc_jpy": 0,
                        "norm_cash": 0, "norm_bank": 0, "norm_jpy": 0,
                        "acad_cash": 0, "acad_bank": 0,
                        "point_qty": 0, "total_qty": 0
                    }

                st = stats[prod_id]
                st["total_qty"] += qty

                if p_type == "point":
                    st["point_qty"] += qty
                elif item_currency == "JPY":
                    if d_type == "student": st["disc_jpy"] += qty
                    else: st["norm_jpy"] += qty
                elif d_type == "student":
                    if p_type == "cash": st["disc_cash"] += qty
                    elif p_type == "bank": st["disc_bank"] += qty
                elif d_type == "academy":
                    if p_type == "cash": st["acad_cash"] += qty
                    elif p_type == "bank": st["acad_bank"] += qty
                else:
                    if p_type == "cash": st["norm_cash"] += qty
                    elif p_type == "bank": st["norm_bank"] += qty

        # 3. 엑셀 워크북 생성 및 스타일 정의
        wb = openpyxl.Workbook()
        ws: Worksheet = wb.active  # type: ignore
        assert ws is not None, "Worksheet 생성 실패"

        ws.title = "일일 판매 보고서"

        font_title = Font(name="맑은 고딕", size=14, bold=True)
        font_header = Font(name="맑은 고딕", size=9, bold=True)
        font_data = Font(name="맑은 고딕", size=9)
        font_sum = Font(name="맑은 고딕", size=10, bold=True)
        
        fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_sub_title = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_sum_row = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        fill_discount_row = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        # 타이틀 레이아웃 (총 20개 컬럼 A~T 적용)
        ws.merge_cells("A1:N2")
        ws["A1"].value = "만물복귀 카페팀 일일 판매 보고서"
        ws["A1"].font = font_title
        ws["A1"].fill = fill_green
        ws["A1"].alignment = align_center

        ws.merge_cells("O1:T2")
        formatted_date = datetime.strptime(target_date, "%Y-%m-%d").strftime("%Y. %m. %d")
        ws["O1"].value = formatted_date
        ws["O1"].font = font_title
        ws["O1"].fill = fill_green
        ws["O1"].alignment = align_center

        for r_num in range(1, 3):
            for c_num in range(1, 21):
                ws.cell(row=r_num, column=c_num).border = thin_border

        # 컬럼 헤더 세팅
        # A~B: 기본정보, C~F: 가격표(확장), G~I: 수련생할인 수량, J~L: 일반 수량, M~N: 아카데미 수량, O~T: 합계/판매량
        ws.merge_cells("A3:A4"); ws.merge_cells("B3:B4")
        ws.merge_cells("C3:F3")  # 가격표 확장 (정가, 엔화정가, 수련생할인, 아카데미할인)
        ws.merge_cells("G3:I3"); ws.merge_cells("J3:L3"); ws.merge_cells("M3:N3")
        ws.merge_cells("O3:O4"); ws.merge_cells("P3:P4"); ws.merge_cells("Q3:Q4")
        ws.merge_cells("R3:R4"); ws.merge_cells("S3:S4"); ws.merge_cells("T3:T4")

        headers_row3 = [
            ("A3", "카테고리"), ("B3", "품목"), ("C3", "가격표"), 
            ("G3", "수련생 할인"), ("J3", "일반"), ("M3", "아카데미 할인"), 
            ("O3", "엔화 합계"), ("P3", "현금 합계"), ("Q3", "계좌 합계"), 
            ("R3", "총 합계"), ("S3", "아카데미 포인트"), ("T3", "총 판매량")
        ]
        for cell_ref, text in headers_row3:
            ws[cell_ref].value = text
            ws[cell_ref].font = font_header
            ws[cell_ref].alignment = align_center
            ws[cell_ref].fill = fill_sub_title

        sub_headers = {
            "C4": "원화 정가", "D4": "엔화 정가", "E4": "수련생 할인액", "F4": "아카데미 할인액",
            "G4": "현금", "H4": "계좌", "I4": "엔화(현금)",
            "J4": "현금", "K4": "계좌", "L4": "엔화(현금)",
            "M4": "현금", "N4": "계좌"
        }
        for cell_ref, text in sub_headers.items():
            ws[cell_ref].value = text
            ws[cell_ref].font = font_header
            ws[cell_ref].alignment = align_center
            ws[cell_ref].fill = fill_sub_title

        for r_num in range(3, 5):
            for c_num in range(1, 21):
                ws.cell(row=r_num, column=c_num).border = thin_border

        # 데이터 세부 행 채우기
        start_data_row = 5
        row_idx = start_data_row

        for prod_id, data in stats.items():
            category = data.get("category", "기타")
            name = data["name"]
            price = data["price"]
            jpy_price = int(round(price / 10))
            disc_student = data.get("discount_student", 0)
            disc_academy = data.get("discount_academy", 0)

            ws.cell(row=row_idx, column=1, value=category)
            ws.cell(row=row_idx, column=2, value=name)
            
            # 가격표 영역 (C~F)
            ws.cell(row=row_idx, column=3, value=price)
            ws.cell(row=row_idx, column=4, value=jpy_price)
            ws.cell(row=row_idx, column=5, value=disc_student)
            ws.cell(row=row_idx, column=6, value=disc_academy)

            # 수량 영역 (G~N)
            ws.cell(row=row_idx, column=7, value=data["disc_cash"] or None)
            ws.cell(row=row_idx, column=8, value=data["disc_bank"] or None)
            ws.cell(row=row_idx, column=9, value=data["disc_jpy"] or None)
            
            ws.cell(row=row_idx, column=10, value=data["norm_cash"] or None)
            ws.cell(row=row_idx, column=11, value=data["norm_bank"] or None)
            ws.cell(row=row_idx, column=12, value=data["norm_jpy"] or None)
            
            ws.cell(row=row_idx, column=13, value=data["acad_cash"] or None)
            ws.cell(row=row_idx, column=14, value=data["acad_bank"] or None)

            # 정가 기준 매출 합계 수식 (O~T)
            r_num = row_idx
            ws.cell(row=r_num, column=15, value=f"=(I{r_num}+L{r_num})*D{r_num}")                  # O열: 엔화 합계 = (I+L)*D(엔화정가)
            ws.cell(row=r_num, column=16, value=f"=(G{r_num}+J{r_num}+M{r_num})*C{r_num}")          # P열: 현금 합계 = (G+J+M)*C(원화정가)
            ws.cell(row=r_num, column=17, value=f"=(H{r_num}+K{r_num}+N{r_num})*C{r_num}")          # Q열: 계좌 합계 = (H+K+N)*C(원화정가)
            ws.cell(row=r_num, column=18, value=f"=P{r_num}+Q{r_num}")                              # R열: 총 합계 = P+Q
            ws.cell(row=r_num, column=19, value=data["point_qty"] or None)                          # S열: 포인트 수량
            ws.cell(row=r_num, column=20, value=f"=SUM(G{r_num}:N{r_num})+S{r_num}")                # T열: 총 판매량

            row_idx += 1

        end_data_row = row_idx - 1

        # 셀 표시 형식 지정
        for r_num in range(start_data_row, row_idx):
            for c_num in range(1, 21):
                cell = ws.cell(row=r_num, column=c_num)
                cell.border = thin_border
                cell.font = font_data

                if c_num in [1, 2]: cell.alignment = align_left
                elif c_num in [3, 5, 6]: cell.alignment = align_right; cell.number_format = '"₩"#,##0'
                elif c_num == 4: cell.alignment = align_right; cell.number_format = '"¥"#,##0'
                elif c_num in range(7, 15) or c_num in [19, 20]: cell.alignment = align_center; cell.number_format = '#,##0'
                elif c_num == 15: cell.alignment = align_right; cell.number_format = '"¥"#,##0'
                else: cell.alignment = align_right; cell.number_format = '"₩"#,##0'

        # 4. 정가 기준 주문 소계 행
        subtotal_row = row_idx
        ws.merge_cells(start_row=subtotal_row, start_column=1, end_row=subtotal_row, end_column=6)
        ws.cell(row=subtotal_row, column=1, value="주문 소계 (정가)").alignment = align_center

        for col_idx in range(7, 21):
            col_letter = get_column_letter(col_idx)
            ws.cell(row=subtotal_row, column=col_idx, value=f"=SUM({col_letter}{start_data_row}:{col_letter}{end_data_row})")

        for c_num in range(1, 21):
            cell = ws.cell(row=subtotal_row, column=c_num)
            cell.border = thin_border; cell.font = font_sum; cell.fill = fill_sum_row
            if c_num in range(7, 15) or c_num in [19, 20]: cell.alignment = align_center; cell.number_format = '#,##0'
            elif c_num == 15: cell.alignment = align_right; cell.number_format = '"¥"#,##0'
            elif c_num in [16, 17, 18]: cell.alignment = align_right; cell.number_format = '"₩"#,##0'

        # 5. 수련생 / 아카데미 / 기타 할인 종류별 차감 행 추가
        disc_start_row = subtotal_row + 1
        
        # 5-1. 수련생 할인 차감
        row_student = disc_start_row
        ws.merge_cells(start_row=row_student, start_column=1, end_row=row_student, end_column=14)
        ws.cell(row=row_student, column=1, value="수련생 할인 차감액").alignment = align_center
        ws.cell(row=row_student, column=15, value=-discount_totals["student_jpy"])
        ws.cell(row=row_student, column=16, value=-discount_totals["student_cash_krw"])  # P열 (현금 차감)
        ws.cell(row=row_student, column=17, value=-discount_totals["student_bank_krw"])  # Q열 (계좌 차감)
        ws.cell(row=row_student, column=18, value=f"=P{row_student}+Q{row_student}")

        # 5-2. 아카데미 할인 차감
        row_academy = disc_start_row + 1
        ws.merge_cells(start_row=row_academy, start_column=1, end_row=row_academy, end_column=14)
        ws.cell(row=row_academy, column=1, value="아카데미 할인 차감액").alignment = align_center
        ws.cell(row=row_academy, column=15, value=-discount_totals["academy_jpy"])
        ws.cell(row=row_academy, column=16, value=-discount_totals["academy_cash_krw"]) # P열 (현금 차감)
        ws.cell(row=row_academy, column=17, value=-discount_totals["academy_bank_krw"]) # Q열 (계좌 차감)
        ws.cell(row=row_academy, column=18, value=f"=P{row_academy}+Q{row_academy}")

        # 5-3. 기타 커스텀 할인 차감
        row_custom = disc_start_row + 2
        ws.merge_cells(start_row=row_custom, start_column=1, end_row=row_custom, end_column=14)
        ws.cell(row=row_custom, column=1, value="기타/커스텀 할인 차감액").alignment = align_center
        ws.cell(row=row_custom, column=15, value=-discount_totals["custom_jpy"])
        ws.cell(row=row_custom, column=16, value=-discount_totals["custom_cash_krw"])    # P열 (현금 차감)
        ws.cell(row=row_custom, column=17, value=-discount_totals["custom_bank_krw"])    # Q열 (계좌 차감)
        ws.cell(row=row_custom, column=18, value=f"=P{row_custom}+Q{row_custom}")

        for r_num in range(disc_start_row, disc_start_row + 3):
            for c_num in range(1, 21):
                cell = ws.cell(row=r_num, column=c_num)
                cell.border = thin_border; cell.font = font_sum; cell.fill = fill_discount_row
                if c_num == 15: cell.alignment = align_right; cell.number_format = '"¥"#,##0'
                elif c_num in [16, 17, 18]: cell.alignment = align_right; cell.number_format = '"₩"#,##0'

        # 6. 최종 실매출 합계 행 (정가 소계 + 할인 차감액)
        final_row = disc_start_row + 3
        ws.merge_cells(start_row=final_row, start_column=1, end_row=final_row, end_column=14)
        ws.cell(row=final_row, column=1, value="최종 실매출 합계").alignment = align_center

        ws.cell(row=final_row, column=15, value=f"=O{subtotal_row}+SUM(O{disc_start_row}:O{disc_start_row+2})")
        ws.cell(row=final_row, column=16, value=f"=P{subtotal_row}+SUM(P{disc_start_row}:P{disc_start_row+2})")
        ws.cell(row=final_row, column=17, value=f"=Q{subtotal_row}+SUM(Q{disc_start_row}:Q{disc_start_row+2})")
        ws.cell(row=final_row, column=18, value=f"=R{subtotal_row}+SUM(R{disc_start_row}:R{disc_start_row+2})")
        ws.cell(row=final_row, column=19, value=f"=S{subtotal_row}")
        ws.cell(row=final_row, column=20, value=f"=T{subtotal_row}")

        for c_num in range(1, 21):
            cell = ws.cell(row=final_row, column=c_num)
            cell.border = thin_border; cell.font = font_sum; cell.fill = fill_green
            if c_num in [19, 20]: cell.alignment = align_center; cell.number_format = '#,##0'
            elif c_num == 15: cell.alignment = align_right; cell.number_format = '"¥"#,##0'
            elif c_num in [16, 17, 18]: cell.alignment = align_right; cell.number_format = '"₩"#,##0'

        wb.save(export_file_path)
        print(f"[Model] 엑셀 보고서 내보내기 성공: {export_file_path}")
        return True