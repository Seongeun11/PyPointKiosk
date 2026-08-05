import os
import shutil
import openpyxl
from typing import List, Dict, Any

class ExcelExporter:
    """
    1안 방식: 카페팀 정산.xlsx 템플릿 서식을 유지하며
    일일 매출 데이터를 정확한 셀 위치에 주입하는 서비스 클래스
    """
    
    # 엑셀 행 매핑 (상품명 -> Excel Row)
    ROW_MAP = {
        "매실": 6, "자허블": 7, "아아": 8, "히비스커스": 9, "밀크티": 10,
        "복숭아 요거트": 11, "딸기라떼": 12, "쇼콜라모카": 13, "초코라떼": 14,
        "말차라떼": 15, "블루베리": 16, "커피젤리": 17, "쫀득쿠키": 18,
        "디아망쿠키": 19, "디아망쿠키 세트": 20, "크림치즈 프레첼": 21,
        "휘낭시에 (3)": 22, "휘낭시에 (5)": 23, "버터떡(2)": 24, "버터떡(4)": 25,
        "티라미수": 26, "바나나 푸딩": 27, "스팸 무스비": 28, "흑임자": 29
    }

    # 결제 수단 & 할인 유형 -> 열(Column) 매핑
    # H:할인가현금(8), I:할인가계좌(9), J:일반가현금(10), K:일반가계좌(11), L:아카데미현금(12), M:아카데미계좌(13), N:엔화(14), S:포인트(19)
    COL_MAP = {
        ("student", "cash"): 8,    # 수련생/할인가 - 현금
        ("student", "bank"): 9,    # 수련생/할인가 - 계좌
        ("none", "cash"): 10,      # 일반가 - 현금
        ("none", "bank"): 11,      # 일반가 - 계좌
        ("academy", "cash"): 12,   # 아카데미 - 현금
        ("academy", "bank"): 13,   # 아카데미 - 계좌
        ("yen", "yen"): 14,        # 엔화
        ("point", "point"): 19     # 아카데미 포인트
    }

    @classmethod
    def export_daily_sales(cls, template_path: str, save_path: str, sales_data: List[Dict[str, Any]], point_history: List[Dict[str, Any]]):
        """
        :param template_path: '카페팀 정산.xlsx' 템플릿 경로
        :param save_path: 내보낼 (.xlsx) 파일 경로
        :param sales_data: 당일 결제 내역 리스트 [ {'name': '아아', 'discount_type': 'none', 'pay_type': 'cash', 'quantity': 2}, ... ]
        :param point_history: 포인트 결제 상세 내역 [ {'name': '김슈야', 'generation': 11, 'point': 3000, 'memo': '히비스커스'}, ... ]
        """
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"엑셀 템플릿 파일을 찾을 수 없습니다: {template_path}")

        # 1. 템플릿 파일 복사
        shutil.copy(template_path, save_path)

        # 2. 복사된 엑셀 파일 로드 (수식 및 서식 유지)
        wb = openpyxl.load_workbook(save_path)
        ws = wb.active
        if ws is None:
            wb.close()
            raise ValueError(f"엑셀 파일에서 활성 워크시트를 찾을 수 없습니다: {save_path}")

        # 3. 판매 데이터 수량 집계 (Matrix)
        qty_matrix = {}  # (row, col) -> total_qty
        
        for sale in sales_data:
            p_name = sale.get("name", "").strip()
            disc_type = sale.get("discount_type", "none") or "none"
            pay_type = sale.get("pay_type", "cash")
            qty = sale.get("quantity", 1)

            # 상품 행 확인
            row_idx = cls.ROW_MAP.get(p_name)
            if not row_idx:
                # 부분 일치 검색
                for k, v in cls.ROW_MAP.items():
                    if k in p_name or p_name in k:
                        row_idx = v
                        break
            
            # 컬럼 확인
            col_idx = cls.COL_MAP.get((disc_type, pay_type))
            if pay_type == "point":
                col_idx = cls.COL_MAP.get(("point", "point"))

            if row_idx and col_idx:
                key = (row_idx, col_idx)
                qty_matrix[key] = qty_matrix.get(key, 0) + qty

        # 4. 수량 셀 주입
        for (row, col), qty in qty_matrix.items():
            cell = ws.cell(row=row, column=col)
            cell.value = (cell.value or 0) + qty

        # 5. 하단 포인트 사용 내역 주입 (Row 37부터)
        if point_history:
            start_row = 37
            for idx, p_info in enumerate(point_history):
                r = start_row + idx
                if r > 44:  # 하단 영역 한계 방지
                    break
                ws.cell(row=r, column=9, value=idx + 1)                      # I열: 순번
                ws.cell(row=r, column=10, value=p_info.get("name", ""))       # J열: 이름
                ws.cell(row=r, column=11, value=p_info.get("generation", "")) # K열: 기수
                ws.cell(row=r, column=12, value=p_info.get("point", 0))       # L열: 사용 포인트
                ws.cell(row=r, column=13, value=p_info.get("memo", ""))        # M열: 사용 내역

        # 6. 저장
        wb.save(save_path)
        wb.close()
        return save_path