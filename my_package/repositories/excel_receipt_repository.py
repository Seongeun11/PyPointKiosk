import os
import json
from datetime import datetime
from typing import Optional, Dict, Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from my_package.repositories.receipt_json_repository import ReceiptRepositoryModel


class ReceiptExcelExporter:
    """
    영수증 데이터 기반 일일 판매 보고서 엑셀 변환 전담 Service
    """
    def __init__(self, receipt_repository: Optional[ReceiptRepositoryModel] = None):
        self.repository = receipt_repository or ReceiptRepositoryModel()

    def _sanitize_excel_text(self, text: str) -> str:
        """엑셀에서 수식(=, +, -)으로 오인하여 시트 파손을 일으키지 않도록 ই스케이프 처리"""
        if not text:
            return ""
        stripped = text.strip()
        # = 나 + 로 시작하면 엑셀이 수식으로 오인하므로 앞에 공백 추가
        if stripped.startswith("=") or stripped.startswith("+"):
            return "'" + text
        return text

    def export_to_excel(self, export_file_path: str, target_date: Optional[str] = None) -> bool:
        if not target_date:
            target_date = self.repository.get_business_date_str()

        default_dir = os.path.join(self.repository.project_root, "resources/reports")
        os.makedirs(default_dir, exist_ok=True)

        if not export_file_path:
            export_file_path = os.path.join(default_dir, f"일일_판매_보고서_{target_date}.xlsx")
        else:
            if not export_file_path.endswith(".xlsx"):
                os.makedirs(export_file_path, exist_ok=True)
                export_file_path = os.path.join(export_file_path, f"일일_판매_보고서_{target_date}.xlsx")
            else:
                base, ext = os.path.splitext(export_file_path)
                if not base.endswith(target_date):
                    export_file_path = f"{base}_{target_date}{ext}"

        receipts = self.repository.get_receipts_by_date(target_date)
        
        # 1. Master 상품 목록 구조화
        stats: Dict[str, Dict[str, Any]] = {}
        if os.path.exists(self.repository.products_path):
            try:
                with open(self.repository.products_path, "r", encoding="utf-8") as f:
                    cat_data = json.load(f).get("categories", [])
                    for cat in cat_data:
                        c_name = cat.get("name", "미지정")
                        for p in cat.get("products", []):
                            prod_id = str(p.get("id"))
                            stats[prod_id] = {
                                "id": prod_id,
                                "name": p.get("name", ""),
                                "category": c_name,
                                "price": p.get("price", 0),
                                "discount_student": p.get("discount_student", 0),
                                "discount_academy": p.get("discount_academy", 0),
                                "disc_cash": 0, "disc_bank": 0, "disc_jpy": 0,
                                "norm_cash": 0, "norm_bank": 0, "norm_jpy": 0,
                                "acad_cash": 0, "acad_bank": 0,
                                "point_qty": 0, "total_qty": 0
                            }
            except Exception as e:
                print(f"[Model Error] 상품 목록 로드 실패: {e}")

        # [핵심 1] 할인 종류별 차감액 & 받은 쿠폰(지불수단) 집계 변수
        discount_totals = {
            "student_cash_krw": 0, "student_bank_krw": 0, "student_jpy": 0,
            "academy_cash_krw": 0, "academy_bank_krw": 0, "academy_jpy": 0,
            "coupon_cash_krw": 0,  "coupon_bank_krw": 0,  "coupon_jpy": 0
        }
        
        # 받은 쿠폰(지불수단) 금액 집계 변수
        coupon_received_totals = {
            "cash_krw": 0,
            "bank_krw": 0,
            "jpy": 0
        }

        # 2. 영수증 JSON 내역 상세 집계
        for r in receipts:
            # [핵심] 취소된 주문일 경우 엑셀 통계 집계에서 제외
            if r.get("is_canceled", False):
                continue

            p_type = r.get("pay_type")        
            d_type = r.get("discount_type")   
            currency = r.get("currency", "KRW")
            r_discount_amt = r.get("discount_amount", 0)
            
            # [신규] JSON 내 받은 쿠폰(지불수단) 금액 집계
            r_coupon_received = r.get("coupon_received", 0)
            if r_coupon_received > 0:
                if currency == "JPY":
                    coupon_received_totals["jpy"] += r_coupon_received
                elif p_type == "bank":
                    coupon_received_totals["bank_krw"] += r_coupon_received
                else:
                    coupon_received_totals["cash_krw"] += r_coupon_received

            items = r.get("items", [])

            # 할인 유형별 차감액 집계
            if d_type == "student":
                if currency == "JPY": 
                    discount_totals["student_jpy"] += r_discount_amt
                elif p_type == "bank": 
                    discount_totals["student_bank_krw"] += r_discount_amt
                else: 
                    discount_totals["student_cash_krw"] += r_discount_amt

            elif d_type == "academy":
                if currency == "JPY": 
                    discount_totals["academy_jpy"] += r_discount_amt
                elif p_type == "bank": 
                    discount_totals["academy_bank_krw"] += r_discount_amt
                else: 
                    discount_totals["academy_cash_krw"] += r_discount_amt

            elif d_type == "coupon" or r_discount_amt > 0:
                if currency == "JPY": 
                    discount_totals["coupon_jpy"] += r_discount_amt
                elif p_type == "bank": 
                    discount_totals["coupon_bank_krw"] += r_discount_amt
                else: 
                    discount_totals["coupon_cash_krw"] += r_discount_amt

            for item in items:
                prod_id = str(item.get("id", item.get("name")))
                name = item.get("name", "미등록상품")
                qty = item.get("quantity", 0)
                item_currency = item.get("currency", currency)

                if prod_id not in stats:
                    stats[prod_id] = {
                        "id": prod_id,
                        "name": name,
                        "category": "기타", 
                        "price": item.get("price", 0),
                        "discount_student": item.get("discount_student", 0),
                        "discount_academy": item.get("discount_academy", 0),
                        "disc_cash": 0, "disc_bank": 0, "disc_jpy": 0,
                        "norm_cash": 0, "norm_bank": 0, "norm_jpy": 0,
                        "acad_cash": 0, "acad_bank": 0,
                        "point_qty": 0, "total_qty": 0
                    }

                st = stats[prod_id]
                st["total_qty"] += qty

                if p_type == "point":
                    st["point_qty"] += qty
                elif item_currency == "JPY":
                    if d_type == "student": st["disc_jpy"] += qty
                    else: st["norm_jpy"] += qty
                elif d_type == "student":
                    if p_type == "cash": st["disc_cash"] += qty
                    elif p_type == "bank": st["disc_bank"] += qty
                elif d_type == "academy":
                    if p_type == "cash": st["acad_cash"] += qty
                    elif p_type == "bank": st["acad_bank"] += qty
                else:
                    if p_type == "cash": st["norm_cash"] += qty
                    elif p_type == "bank": st["norm_bank"] += qty

        # 3. 엑셀 워크북 생성 및 스타일 정의
        wb = openpyxl.Workbook()
        ws: Worksheet = wb.active  # type: ignore
        ws.title = "일일 판매 보고서"

        font_title = Font(name="맑은 고딕", size=14, bold=True)
        font_header = Font(name="맑은 고딕", size=9, bold=True)
        font_data = Font(name="맑은 고딕", size=9)
        font_sum = Font(name="맑은 고딕", size=10, bold=True)
        
        fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_sub_title = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_sum_row = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        fill_discount_row = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        fill_coupon_pay_row = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        # 타이틀 레이아웃
        ws.merge_cells("A1:N2")
        ws["A1"].value = "만물복귀 카페팀 일일 판매 보고서"
        ws["A1"].font = font_title; ws["A1"].fill = fill_green; ws["A1"].alignment = align_center

        ws.merge_cells("O1:T2")
        formatted_date = datetime.strptime(target_date, "%Y-%m-%d").strftime("%Y. %m. %d")
        ws["O1"].value = formatted_date
        ws["O1"].font = font_title; ws["O1"].fill = fill_green; ws["O1"].alignment = align_center

        for r_num in range(1, 3):
            for c_num in range(1, 21):
                ws.cell(row=r_num, column=c_num).border = thin_border

        # 헤더 세팅
        ws.merge_cells("A3:A4"); ws.merge_cells("B3:B4"); ws.merge_cells("C3:F3")
        ws.merge_cells("G3:I3"); ws.merge_cells("J3:L3"); ws.merge_cells("M3:N3")
        ws.merge_cells("O3:O4"); ws.merge_cells("P3:P4"); ws.merge_cells("Q3:Q4")
        ws.merge_cells("R3:R4"); ws.merge_cells("S3:S4"); ws.merge_cells("T3:T4")

        headers_row3 = [
            ("A3", "카테고리"), ("B3", "품목"), ("C3", "가격표"), 
            ("G3", "수련생"), ("J3", "일반"), ("M3", "아카데미"), 
            ("O3", "엔화 합계"), ("P3", "현금 합계"), ("Q3", "계좌 합계"), 
            ("R3", "총 합계"), ("S3", "아카데미 포인트"), ("T3", "총 판매량")
        ]
        for cell_ref, text in headers_row3:
            ws[cell_ref].value = text
            ws[cell_ref].font = font_header; ws[cell_ref].alignment = align_center; ws[cell_ref].fill = fill_sub_title

        sub_headers = {
            "C4": "원화 정가", "D4": "엔화 정가", "E4": "수련생 할인액", "F4": "아카데미 할인액",
            "G4": "현금", "H4": "계좌", "I4": "엔화(현금)",
            "J4": "현금", "K4": "계좌", "L4": "엔화(현금)",
            "M4": "현금", "N4": "계좌"
        }
        for cell_ref, text in sub_headers.items():
            ws[cell_ref].value = text
            ws[cell_ref].font = font_header; ws[cell_ref].alignment = align_center; ws[cell_ref].fill = fill_sub_title

        for r_num in range(3, 5):
            for c_num in range(1, 21):
                ws.cell(row=r_num, column=c_num).border = thin_border

        # 데이터 행 작성
        start_data_row = 5
        row_idx = start_data_row

        for prod_id, data in stats.items():
            category = data.get("category", "기타")
            name = data["name"]
            price = data["price"]
            jpy_price = int(round(price / 10))
            disc_student = data.get("discount_student", 0)
            disc_academy = data.get("discount_academy", 0)

            ws.cell(row=row_idx, column=1, value=category)
            ws.cell(row=row_idx, column=2, value=name)
            ws.cell(row=row_idx, column=3, value=price)
            ws.cell(row=row_idx, column=4, value=jpy_price)
            ws.cell(row=row_idx, column=5, value=disc_student)
            ws.cell(row=row_idx, column=6, value=disc_academy)

            ws.cell(row=row_idx, column=7, value=data["disc_cash"] or None)
            ws.cell(row=row_idx, column=8, value=data["disc_bank"] or None)
            ws.cell(row=row_idx, column=9, value=data["disc_jpy"] or None)
            
            ws.cell(row=row_idx, column=10, value=data["norm_cash"] or None)
            ws.cell(row=row_idx, column=11, value=data["norm_bank"] or None)
            ws.cell(row=row_idx, column=12, value=data["norm_jpy"] or None)
            
            ws.cell(row=row_idx, column=13, value=data["acad_cash"] or None)
            ws.cell(row=row_idx, column=14, value=data["acad_bank"] or None)

            r_num = row_idx
            ws.cell(row=r_num, column=15, value=f"=(I{r_num}+L{r_num})*D{r_num}")
            ws.cell(row=r_num, column=16, value=f"=(G{r_num}+J{r_num}+M{r_num})*C{r_num}")
            ws.cell(row=r_num, column=17, value=f"=(H{r_num}+K{r_num}+N{r_num})*C{r_num}")
            ws.cell(row=r_num, column=18, value=f"=P{r_num}+Q{r_num}")
            ws.cell(row=r_num, column=19, value=data["point_qty"] or None)
            ws.cell(row=r_num, column=20, value=f"=SUM(G{r_num}:N{r_num})+S{r_num}")

            row_idx += 1

        end_data_row = row_idx - 1

        # 데이터 셀 서식 적용
        for r_num in range(start_data_row, row_idx):
            for c_num in range(1, 21):
                cell = ws.cell(row=r_num, column=c_num)
                cell.border = thin_border; cell.font = font_data
                if c_num in [1, 2]: cell.alignment = align_left
                elif c_num in [3, 5, 6]: cell.alignment = align_right; cell.number_format = '"₩"#,##0'
                elif c_num == 4: cell.alignment = align_right; cell.number_format = '"¥"#,##0'
                elif c_num in range(7, 15) or c_num in [19, 20]: cell.alignment = align_center; cell.number_format = '#,##0'
                elif c_num == 15: cell.alignment = align_right; cell.number_format = '"¥"#,##0'
                else: cell.alignment = align_right; cell.number_format = '"₩"#,##0'

        # 4. 정가 기준 주문 소계 행
        subtotal_row = row_idx
        ws.merge_cells(start_row=subtotal_row, start_column=1, end_row=subtotal_row, end_column=6)
        ws.cell(row=subtotal_row, column=1, value="주문 소계 (정가)").alignment = align_center

        for col_idx in range(7, 21):
            col_letter = get_column_letter(col_idx)
            ws.cell(row=subtotal_row, column=col_idx, value=f"=SUM({col_letter}{start_data_row}:{col_letter}{end_data_row})")

        for c_num in range(1, 21):
            cell = ws.cell(row=subtotal_row, column=c_num)
            cell.border = thin_border; cell.font = font_sum; cell.fill = fill_sum_row
            if c_num in range(7, 15) or c_num in [19, 20]: cell.alignment = align_center; cell.number_format = '#,##0'
            elif c_num == 15: cell.alignment = align_right; cell.number_format = '"¥"#,##0'
            elif c_num in [16, 17, 18]: cell.alignment = align_right; cell.number_format = '"₩"#,##0'

        # 5. 할인 종류별 차감 행 (3종)
        disc_start_row = subtotal_row + 1
        
        # 5-1. 수련생 할인 차감
        row_student = disc_start_row
        ws.merge_cells(start_row=row_student, start_column=1, end_row=row_student, end_column=14)
        ws.cell(row=row_student, column=1, value="수련생 할인 차감액").alignment = align_center
        ws.cell(row=row_student, column=15, value=-discount_totals["student_jpy"])
        ws.cell(row=row_student, column=16, value=-discount_totals["student_cash_krw"])
        ws.cell(row=row_student, column=17, value=-discount_totals["student_bank_krw"])
        ws.cell(row=row_student, column=18, value=f"=P{row_student}+Q{row_student}")

        # 5-2. 아카데미 할인 차감
        row_academy = disc_start_row + 1
        ws.merge_cells(start_row=row_academy, start_column=1, end_row=row_academy, end_column=14)
        ws.cell(row=row_academy, column=1, value="아카데미 할인 차감액").alignment = align_center
        ws.cell(row=row_academy, column=15, value=-discount_totals["academy_jpy"])
        ws.cell(row=row_academy, column=16, value=-discount_totals["academy_cash_krw"])
        ws.cell(row=row_academy, column=17, value=-discount_totals["academy_bank_krw"])
        ws.cell(row=row_academy, column=18, value=f"=P{row_academy}+Q{row_academy}")

        # 5-3. 쿠폰 할인 차감
        row_coupon = disc_start_row + 2
        ws.merge_cells(start_row=row_coupon, start_column=1, end_row=row_coupon, end_column=14)
        ws.cell(row=row_coupon, column=1, value="쿠폰 할인 차감액").alignment = align_center
        ws.cell(row=row_coupon, column=15, value=-discount_totals["coupon_jpy"])
        ws.cell(row=row_coupon, column=16, value=-discount_totals["coupon_cash_krw"])
        ws.cell(row=row_coupon, column=17, value=-discount_totals["coupon_bank_krw"])
        ws.cell(row=row_coupon, column=18, value=f"=P{row_coupon}+Q{row_coupon}")

        for r_num in range(disc_start_row, disc_start_row + 3):
            for c_num in range(1, 21):
                cell = ws.cell(row=r_num, column=c_num)
                cell.border = thin_border; cell.font = font_sum; cell.fill = fill_discount_row
                if c_num == 15: cell.alignment = align_right; cell.number_format = '"¥"#,##0'
                elif c_num in [16, 17, 18]: cell.alignment = align_right; cell.number_format = '"₩"#,##0'

        # [핵심 2] 5-4. 받은 쿠폰(지불수단) 차감액 영역 생성 (쿠폰 할인 차감액 영역 바로 밑)
        row_received_coupon = disc_start_row + 3
        ws.merge_cells(start_row=row_received_coupon, start_column=1, end_row=row_received_coupon, end_column=14)
        ws.cell(row=row_received_coupon, column=1, value="쿠폰(지불수단) 차감액").alignment = align_center
        
        ws.cell(row=row_received_coupon, column=15, value=-coupon_received_totals["jpy"])
        ws.cell(row=row_received_coupon, column=16, value=-coupon_received_totals["cash_krw"])
        ws.cell(row=row_received_coupon, column=17, value=-coupon_received_totals["bank_krw"])
        ws.cell(row=row_received_coupon, column=18, value=f"=P{row_received_coupon}+Q{row_received_coupon}")

        for c_num in range(1, 21):
            cell = ws.cell(row=row_received_coupon, column=c_num)
            cell.border = thin_border; cell.font = font_sum; cell.fill = fill_coupon_pay_row
            if c_num == 15: cell.alignment = align_right; cell.number_format = '"¥"#,##0'
            elif c_num in [16, 17, 18]: cell.alignment = align_right; cell.number_format = '"₩"#,##0'

        # [핵심 3] 6. 최종 실매출 합계 행 (주문 소계 + 할인 차감액 3종 + 받은 쿠폰 지불수단 차감액 합산)
        final_row = disc_start_row + 4
        ws.merge_cells(start_row=final_row, start_column=1, end_row=final_row, end_column=14)
        ws.cell(row=final_row, column=1, value="최종 실매출 합계").alignment = align_center
        
        # O, P, Q, R 열에 주문 소계부터 할인 차감 및 받은 쿠폰 차감액까지 전체 합산하는 엑셀 수식 적용
        disc_end_row = row_received_coupon
        
        ws.cell(row=final_row, column=15, value=f"=O{subtotal_row}+SUM(O{disc_start_row}:O{disc_end_row})")
        ws.cell(row=final_row, column=16, value=f"=P{subtotal_row}+SUM(P{disc_start_row}:P{disc_end_row})")
        ws.cell(row=final_row, column=17, value=f"=Q{subtotal_row}+SUM(Q{disc_start_row}:Q{disc_end_row})")
        ws.cell(row=final_row, column=18, value=f"=R{subtotal_row}+SUM(R{disc_start_row}:R{disc_end_row})")
        ws.cell(row=final_row, column=19, value=f"=S{subtotal_row}")
        ws.cell(row=final_row, column=20, value=f"=T{subtotal_row}")

        for c_num in range(1, 21):
            cell = ws.cell(row=final_row, column=c_num)
            cell.border = thin_border; cell.font = font_sum; cell.fill = fill_green
            if c_num in [19, 20]: cell.alignment = align_center; cell.number_format = '#,##0'
            elif c_num == 15: cell.alignment = align_right; cell.number_format = '"¥"#,##0'
            elif c_num in [16, 17, 18]: cell.alignment = align_right; cell.number_format = '"₩"#,##0'
        # --------------------------------------------------------------------------
        # 시트 2: 영수증 내역 (전체 외곽/내부 테두리 및 카드 양식 적용)
        # --------------------------------------------------------------------------
        ws2: Worksheet = wb.create_sheet(title="영수증 내역")
        ws2.views.sheetView[0].showGridLines = True

        # 기본 일반 영수증 스타일[cite: 1]
        font_rcpt_title = Font(name="Consolas", size=11, bold=True, color="0026FF")
        fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # [신규] 취소 영수증 전용 스타일 정의[cite: 1]
        font_rcpt_title_canceled = Font(name="Consolas", size=11, bold=True, color="C00000")
        fill_header_canceled = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        font_rcpt_text = Font(name="Consolas", size=9.5)

        # 테두리 스타일 정의[cite: 1]
        side_outer_thin = Side(style="thin", color="888888")
        side_inner_gray = Side(style="thin", color="E0E0E0")

        curr_row = 2

        if not receipts:
            ws2.cell(
                row=curr_row, column=2, value="해당 일자의 영수증 내역이 없습니다."
            )
        else:
            for idx, r in enumerate(receipts, start=1):
                rcpt_id = r.get("id", idx)
                rcpt_text = r.get("receipt_text", "")
                timestamp = r.get("timestamp", "")
                
                # [핵심] 주문 취소 여부 판단 플래그[cite: 1]
                is_canceled = r.get("is_canceled", False)

                block_start_row = curr_row

                # 1. 영수증 헤더 타이틀 (취소 여부에 따라 텍스트 및 스타일 구분 적용)[cite: 1]
                ws2.merge_cells(
                    start_row=curr_row, start_column=2, end_row=curr_row, end_column=6
                )
                
                # [핵심] 취소 여부에 따른 제목 문구 설정[cite: 1]
                title_prefix = "■ [주문취소]" if is_canceled else "■"
                header_text = f"{title_prefix} 영수증 No. {rcpt_id} ({timestamp})"
                
                header_cell = ws2.cell(
                    row=curr_row,
                    column=2,
                    value=header_text,
                )
                
                # [핵심] 취소 여부에 따른 폰트 지정[cite: 1]
                header_cell.font = font_rcpt_title_canceled if is_canceled else font_rcpt_title
                header_cell.alignment = Alignment(vertical="center")
                header_cell.number_format = "@"

                # [핵심] 헤더 행 B~F 전체에 취소 여부에 따른 배경색 지정[cite: 1]
                current_fill = fill_header_canceled if is_canceled else fill_header
                for c in range(2, 7):
                    ws2.cell(row=curr_row, column=c).fill = current_fill

                curr_row += 1

                # 2. 영수증 텍스트 행 출력[cite: 1]
                if rcpt_text:
                    lines = rcpt_text.splitlines()
                    for line in lines:
                        ws2.merge_cells(
                            start_row=curr_row,
                            start_column=2,
                            end_row=curr_row,
                            end_column=6,
                        )

                        safe_text = self._sanitize_excel_text(line)

                        cell = ws2.cell(row=curr_row, column=2, value=safe_text)
                        cell.font = font_rcpt_text
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.number_format = "@"

                        if "===" in line or "---" in line:
                            cell.font = Font(name="Consolas", size=9, color="888888")

                        curr_row += 1

                    block_end_row = curr_row - 1

                    # 3. 영수증 전체 카드 박스 테두리(Border) 적용 (B~F열, block_start_row ~ block_end_row)[cite: 1]
                    for r_i in range(block_start_row, block_end_row + 1):
                        for c_i in range(2, 7):
                            cell = ws2.cell(row=r_i, column=c_i)

                            # 상, 하, 좌, 우 위치 판단 후 외곽선 및 내부 테두리 지정[cite: 1]
                            top_border = (
                                side_outer_thin if r_i == block_start_row else side_inner_gray
                            )
                            bottom_border = (
                                side_outer_thin if r_i == block_end_row else side_inner_gray
                            )
                            left_border = side_outer_thin if c_i == 2 else Side(style=None)
                            right_border = side_outer_thin if c_i == 6 else Side(style=None)

                            cell.border = Border(
                                top=top_border,
                                bottom=bottom_border,
                                left=left_border,
                                right=right_border,
                            )

                    curr_row += 2  # 영수증 간 간격[cite: 1]

        ws2.column_dimensions["A"].width = 3
        ws2.column_dimensions["B"].width = 55
        wb.save(export_file_path)
        print(f"[Service] 엑셀 보고서 내보내기 성공: {export_file_path}")
        return True